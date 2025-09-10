import re
import csv
import io
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Optional, Tuple

import xlrd
import openpyxl
from django.db import transaction

from .models import Skin
from register.models import Player
from events.models import Event
from courses.models import Course, Hole


class SkinImportError(Exception):
    """Custom exception for skin import errors"""
    pass


class WednesdaySkinsImporter:
    """
    Imports Wednesday individual skins data from Golf Genius leaderboard exports.
    
    Two-step process:
    1. Parse Excel file and generate CSV data
    2. Import CSV data to create Skin records
    """
    
    def __init__(self):
        self.processed_count = 0
        self.skipped_count = 0
    
    def generate_csv_from_excel(self, file_path: str) -> Tuple[str, List[str]]:
        """
        Parse Excel file and generate CSV data for skins import.
        
        Args:
            file_path: Path to the Golf Genius leaderboard Excel file
            
        Returns:
            Tuple of (csv_data_string, list_of_errors)
        """
        failures = []
        csv_rows = []
        
        # Add CSV header
        csv_rows.append([
            'player_name',
            'player_ghin', 
            'course',
            'hole_number',
            'skin_type',
            'value',
            'details'
        ])
        
        try:
            # Parse Excel file
            if file_path.lower().endswith('.xls'):
                sheets_data = self._parse_xls_file(file_path)
            elif file_path.lower().endswith('.xlsx'):
                sheets_data = self._parse_xlsx_file(file_path)
            else:
                failures.append("Unsupported file format. Only .xls and .xlsx files are supported.")
                return "", failures
        except Exception as e:
            failures.append(f"Error reading Excel file: {str(e)}")
            return "", failures
        
        # Process each relevant sheet
        for sheet_name, sheet_data in sheets_data.items():
            if self._is_skins_sheet(sheet_name):
                course_name = self._extract_course_name(sheet_name)
                skin_type = self._determine_skin_type(sheet_name)
                
                sheet_failures, sheet_rows = self._process_skins_sheet_for_csv(
                    sheet_data, course_name, skin_type
                )
                failures.extend(sheet_failures)
                csv_rows.extend(sheet_rows)
        
        # Convert to CSV string
        output = io.StringIO()
        writer = csv.writer(output)
        for row in csv_rows:
            writer.writerow(row)
        csv_data = output.getvalue()
        output.close()
        
        return csv_data, failures
    
    def import_csv_data(self, event: Event, csv_data: str) -> List[str]:
        """
        Import skins data from CSV string and create Skin records.
        
        Args:
            event: The Event these skins belong to
            csv_data: CSV data string with skins information
            
        Returns:
            List of error messages (empty list if successful)
        """
        failures = []
        self.processed_count = 0
        self.skipped_count = 0
        
        # Get existing skins for this event to make import idempotent
        existing_skins = {}
        for skin in Skin.objects.filter(event=event):
            key = (skin.player.id, skin.hole.hole_number, skin.skin_type)
            existing_skins[key] = skin
        
        try:
            # Parse CSV data
            csv_reader = csv.DictReader(io.StringIO(csv_data))
            
            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 for header
                    try:
                        self._process_csv_row(row, event, existing_skins)
                    except Exception as e:
                        failures.append(f"Error processing CSV row {row_num}: {str(e)}")
                        self.skipped_count += 1
                        
        except Exception as e:
            failures.append(f"Error parsing CSV data: {str(e)}")
        
        return failures
    
    def _parse_xls_file(self, file_path: str) -> Dict:
        """Parse .xls file and return sheet data"""
        try:
            workbook = xlrd.open_workbook(file_path)
            sheets_data = {}
            
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_data = []
                
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_data.append(cell_value)
                    sheet_data.append(row_data)
                
                sheets_data[sheet_name] = sheet_data
            
            return sheets_data
            
        except Exception as e:
            raise SkinImportError(f"Error reading .xls file: {str(e)}")
    
    def _parse_xlsx_file(self, file_path: str) -> Dict:
        """Parse .xlsx file and return sheet data"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                
                sheets_data[sheet_name] = sheet_data
            
            return sheets_data
            
        except Exception as e:
            raise SkinImportError(f"Error reading .xlsx file: {str(e)}")
    
    def _is_skins_sheet(self, sheet_name: str) -> bool:
        """Check if sheet name indicates it contains skins data"""
        sheet_name_lower = sheet_name.lower().strip()
        return (sheet_name_lower.startswith('gross skins') or 
                sheet_name_lower.startswith('net skins'))
    
    def _extract_course_name(self, sheet_name: str) -> str:
        """Extract course name from sheet name (east, north, west)"""
        sheet_name_lower = sheet_name.lower()
        
        if 'east' in sheet_name_lower:
            return 'east'
        elif 'north' in sheet_name_lower:
            return 'north'
        elif 'west' in sheet_name_lower:
            return 'west'
        else:
            return 'unknown'
    
    def _determine_skin_type(self, sheet_name: str) -> str:
        """Determine skin type from sheet name"""
        sheet_name_lower = sheet_name.lower().strip()
        if sheet_name_lower.startswith('gross skins'):
            return 'Gross'
        elif sheet_name_lower.startswith('net skins'):
            return 'Net'
        else:
            return 'Gross'  # Default fallback
    
    def _process_skins_sheet_for_csv(self, sheet_data: List[List], 
                                   course_name: str, skin_type: str) -> Tuple[List[str], List[List]]:
        """Process a skins sheet and generate CSV rows"""
        failures = []
        csv_rows = []
        
        if not sheet_data:
            failures.append(f"Empty sheet found for {skin_type} skins - {course_name}")
            return failures, csv_rows
        
        # Find header row and column mapping
        header_row_idx, column_mapping = self._find_skins_header_row(sheet_data)
        
        if header_row_idx is None:
            failures.append(f"Could not find header row in {skin_type} skins sheet - {course_name}")
            return failures, csv_rows
        
        current_flight = None
        
        # Process each data row
        for row_idx in range(header_row_idx + 1, len(sheet_data)):
            row_data = sheet_data[row_idx]
            
            if self._is_empty_row(row_data):
                continue
            
            # Check if this is a flight header row
            flight_name = self._extract_flight_name(row_data)
            if flight_name:
                current_flight = flight_name
                continue
            
            try:
                # Process player row and generate CSV rows for each skin
                player_csv_rows = self._process_player_row_for_csv(
                    row_data, column_mapping, course_name, current_flight, skin_type
                )
                csv_rows.extend(player_csv_rows)
                
            except Exception as e:
                failures.append(f"Error processing row {row_idx + 1} in {course_name} {skin_type}: {str(e)}")
        
        return failures, csv_rows
    
    def _find_skins_header_row(self, sheet_data: List[List]) -> Tuple[Optional[int], Dict]:
        """Find the header row for skins data"""
        player_patterns = ['player']
        skins_patterns = ['skins']
        purse_patterns = ['purse']
        details_patterns = ['details']
        
        for row_idx, row_data in enumerate(sheet_data):
            if not row_data:
                continue
                
            # Convert row to lowercase strings for pattern matching
            row_lower = [str(cell).lower().strip() if cell else '' for cell in row_data]
            
            # Check if this looks like a skins header row
            player_col = self._find_column_index(row_lower, player_patterns)
            skins_col = self._find_column_index(row_lower, skins_patterns)
            purse_col = self._find_column_index(row_lower, purse_patterns)
            details_col = self._find_column_index(row_lower, details_patterns)
            
            if (player_col is not None and skins_col is not None and 
                purse_col is not None and details_col is not None):
                
                return row_idx, {
                    'player': player_col,
                    'skins': skins_col,
                    'purse': purse_col,
                    'details': details_col
                }
        
        return None, {}
    
    def _find_column_index(self, row_data: List[str], patterns: List[str]) -> Optional[int]:
        """Find column index that matches any of the given patterns"""
        for idx, cell_value in enumerate(row_data):
            for pattern in patterns:
                if pattern in cell_value:
                    return idx
        return None
    
    def _is_empty_row(self, row_data: List) -> bool:
        """Check if a row is empty or contains only whitespace"""
        return not any(str(cell).strip() for cell in row_data if cell is not None)
    
    def _extract_flight_name(self, row_data: List) -> Optional[str]:
        """Extract flight name from a row (e.g., 'Flight 1')"""
        for cell in row_data:
            if cell and isinstance(cell, str):
                cell_lower = cell.lower().strip()
                if cell_lower.startswith('flight '):
                    return cell.strip()
        return None
    
    def _process_player_row_for_csv(self, row_data: List, column_mapping: Dict, 
                                  course_name: str, flight: str, skin_type: str) -> List[List]:
        """Process a player row and generate CSV rows for each skin won"""
        csv_rows = []
        
        # Extract player name
        player_name = self._extract_player_name(row_data, column_mapping)
        if not player_name:
            raise ValueError("No player name found")
        
        # Find player in database to get GHIN (but don't fail if not found)
        player = self._find_player(player_name)
        player_ghin = player.ghin if player else ''
        
        # Extract skins count and total purse
        skins_count = self._extract_skins_count(row_data, column_mapping)
        total_purse = self._extract_total_purse(row_data, column_mapping)
        details = self._extract_details(row_data, column_mapping)
        
        if skins_count == 0:
            return csv_rows
        
        # Parse individual skins from details
        individual_skins = self._parse_skin_details(details, skins_count, total_purse, flight)
        
        # Create CSV rows for each skin
        for skin_info in individual_skins:
            csv_rows.append([
                player_name,
                player_ghin,
                course_name,
                skin_info['hole_number'],
                skin_type,
                str(skin_info['value']),
                skin_info['details']
            ])
        
        return csv_rows
    
    def _extract_player_name(self, row_data: List, column_mapping: Dict) -> Optional[str]:
        """Extract and clean player name from row data"""
        player_col = column_mapping.get('player')
        if player_col is None or player_col >= len(row_data):
            return None
        
        player_name = str(row_data[player_col]).strip()
        return player_name if player_name and player_name.lower() != 'none' else None
    
    def _find_player(self, player_name: str) -> Optional[Player]:
        """Find player by name (only includes members - is_member=True)"""
        try:
            # Try "Last, First" format
            if ',' in player_name:
                last_name, first_name = [name.strip() for name in player_name.split(',', 1)]
                return Player.objects.get(
                    first_name__iexact=first_name,
                    last_name__iexact=last_name,
                    is_member=True
                )
            
            # Try "First Last" format
            elif ' ' in player_name:
                name_parts = player_name.split()
                if len(name_parts) >= 2:
                    first_name = name_parts[0]
                    last_name = ' '.join(name_parts[1:])
                    return Player.objects.get(
                        first_name__iexact=first_name,
                        last_name__iexact=last_name,
                        is_member=True
                    )
                    
        except (Player.DoesNotExist, Player.MultipleObjectsReturned):
            pass
        
        return None
    
    def _extract_skins_count(self, row_data: List, column_mapping: Dict) -> int:
        """Extract number of skins won from row data"""
        skins_col = column_mapping.get('skins')
        if skins_col is None or skins_col >= len(row_data):
            return 0
        
        skins_value = row_data[skins_col]
        try:
            return int(skins_value) if skins_value else 0
        except (ValueError, TypeError):
            return 0
    
    def _extract_total_purse(self, row_data: List, column_mapping: Dict) -> Decimal:
        """Extract total purse amount from row data"""
        purse_col = column_mapping.get('purse')
        if purse_col is None or purse_col >= len(row_data):
            return Decimal('0.00')
        
        purse_value = row_data[purse_col]
        if not purse_value:
            return Decimal('0.00')
        
        try:
            # Clean up the purse string (remove $ and other non-numeric characters)
            purse_str = str(purse_value).strip()
            purse_str = re.sub(r'[^\d.]', '', purse_str)
            return Decimal(purse_str) if purse_str else Decimal('0.00')
        except (InvalidOperation, ValueError):
            return Decimal('0.00')
    
    def _extract_details(self, row_data: List, column_mapping: Dict) -> str:
        """Extract details string from row data"""
        details_col = column_mapping.get('details')
        if details_col is None or details_col >= len(row_data):
            return ""
        
        details_value = row_data[details_col]
        return str(details_value).strip() if details_value else ""
    
    def _parse_skin_details(self, details: str, skins_count: int, 
                          total_purse: Decimal, flight: str) -> List[Dict]:
        """
        Parse individual skin details from the details string.
        
        Examples: 
        - "Birdie on 7" -> [{'hole_number': 7, 'value': 38.00, 'details': 'Birdie on 7 in Flight 1'}]
        - "Birdie on 4, Par on 7" -> [{'hole_number': 4, ...}, {'hole_number': 7, ...}]
        """
        individual_skins = []
        
        if not details or skins_count == 0:
            return individual_skins
        
        # Calculate individual skin value
        individual_value = total_purse / skins_count if skins_count > 0 else Decimal('0.00')
        
        # Split details by comma and parse each skin
        skin_details = [detail.strip() for detail in details.split(',')]
        
        for detail in skin_details:
            hole_number = self._extract_hole_from_detail(detail)
            if hole_number:
                full_detail = f"{detail} in {flight}" if flight else detail
                individual_skins.append({
                    'hole_number': hole_number,
                    'value': individual_value,
                    'details': full_detail
                })
        
        # If we couldn't parse individual holes, create generic entries
        if not individual_skins and skins_count > 0:
            for i in range(skins_count):
                individual_skins.append({
                    'hole_number': 1,  # Default to hole 1
                    'value': individual_value,
                    'details': f"{details} in {flight}" if flight else details
                })
        
        return individual_skins
    
    def _extract_hole_from_detail(self, detail: str) -> Optional[int]:
        """Extract hole number from detail string like 'Birdie on 7'"""
        # Look for patterns like "on 7", "hole 7", etc.
        match = re.search(r'\bon\s+(\d+)', detail, re.IGNORECASE)
        if match:
            hole_num = int(match.group(1))
            return hole_num if 1 <= hole_num <= 18 else None
        
        # Look for just numbers at the end
        match = re.search(r'\b(\d+)\b', detail)
        if match:
            hole_num = int(match.group(1))
            return hole_num if 1 <= hole_num <= 18 else None
        
        return None
    
    def _process_csv_row(self, row: Dict, event: Event, existing_skins: Dict):
        """Process a single CSV row and create/update Skin record"""
        # Extract data from CSV row
        player_name = row['player_name'].strip()
        course_name = row['course'].strip().lower()
        hole_number = int(row['hole_number'])
        skin_type = row.get('skin_type', 'Gross').strip()
        value = Decimal(row['value'])
        details = row['details'].strip()
        
        # Find player
        player = self._find_player(player_name)
        if not player:
            raise ValueError(f"Player '{player_name}' not found")
        
        # Find or create course
        course = self._find_or_create_course(course_name)
        
        # Find or create hole
        hole = self._find_or_create_hole(course, hole_number)
        
        # Check if this skin already exists (idempotent operation)
        skin_key = (player.id, hole_number, skin_type)
        existing_skin = existing_skins.get(skin_key)
        
        if existing_skin:
            # Update existing record if the amount has changed
            if existing_skin.skin_amount != value or existing_skin.details != details:
                existing_skin.skin_amount = value
                existing_skin.details = details
                existing_skin.save()
                self.processed_count += 1
        else:
            # Create new skin record
            skin = Skin.objects.create(
                event=event,
                course=course,
                hole=hole,
                player=player,
                skin_type=skin_type,
                skin_amount=value,
                is_team=False,
                details=details
            )
            # Add to existing_skins to prevent duplicates in same import
            existing_skins[skin_key] = skin
            self.processed_count += 1
    
    def _find_or_create_course(self, course_name: str) -> Course:
        """Find or create course by name"""
        course_map = {
            'east': 'Bunker Hills East',
            'north': 'Bunker Hills North', 
            'west': 'Bunker Hills West'
        }
        
        full_course_name = course_map.get(course_name, f"Bunker Hills {course_name.title()}")
        
        course, created = Course.objects.get_or_create(
            name=full_course_name,
            defaults={'number_of_holes': 18}
        )
        return course
    
    def _find_or_create_hole(self, course: Course, hole_number: int) -> Hole:
        """Find or create hole for the given course and number"""
        hole, created = Hole.objects.get_or_create(
            course=course,
            hole_number=hole_number,
            defaults={'par': 4}  # Default par
        )
        return hole


# Convenience functions
def generate_skins_csv_from_excel(file_path: str) -> Tuple[str, List[str]]:
    """
    Generate CSV data from Excel file.
    
    Args:
        file_path: Path to the Golf Genius leaderboard Excel file
        
    Returns:
        Tuple of (csv_data_string, list_of_errors)
    """
    importer = WednesdaySkinsImporter()
    return importer.generate_csv_from_excel(file_path)


def import_skins_from_csv(event: Event, csv_data: str) -> List[str]:
    """
    Import skins from CSV data.
    
    Args:
        event: The Event these skins belong to
        csv_data: CSV data string with skins information
        
    Returns:
        List of error messages (empty list if successful)
    """
    importer = WednesdaySkinsImporter()
    return importer.import_csv_data(event, csv_data)
