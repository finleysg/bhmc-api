import random
import string

from datetime import timedelta

import djoser.views
from djoser import utils
from djoser.conf import settings

from openpyxl import load_workbook

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import permission_classes, api_view
from rest_framework.response import Response

from bhmc.settings import is_development, to_bool
from documents.models import Document
from events.models import Event
from register.models import Player
from .models import BoardMember, MajorChampion, Ace, LowScore, SeasonSettings
from .serializers import BoardMemberSerializer, AceSerializer, LowScoreSerializer, MajorChampionSerializer, \
    SeasonSettingsSerializer


is_localhost = to_bool(is_development)

class BoardMemberViewSet(viewsets.ModelViewSet):
    serializer_class = BoardMemberSerializer
    queryset = BoardMember.objects.all()


class MajorChampionViewSet(viewsets.ModelViewSet):
    serializer_class = MajorChampionSerializer

    def get_queryset(self):
        queryset = MajorChampion.objects.all()
        season = self.request.query_params.get("season", None)
        event_id = self.request.query_params.get("event", None)
        player_id = self.request.query_params.get("player", None)

        if season is not None:
            queryset = queryset.filter(season=season)
        if event_id is not None:
            queryset = queryset.filter(event=event_id)
        if player_id is not None:
            queryset = queryset.filter(player=player_id)
        queryset = queryset.order_by("-season", "event_name", "flight")

        return queryset


class LowScoreViewSet(viewsets.ModelViewSet):
    serializer_class = LowScoreSerializer

    def get_queryset(self):
        queryset = LowScore.objects.all()
        season = self.request.query_params.get("season", None)

        if season is not None:
            queryset = queryset.filter(season=season)

        return queryset


class AceViewSet(viewsets.ModelViewSet):
    serializer_class = AceSerializer

    def get_queryset(self):
        queryset = Ace.objects.all()
        season = self.request.query_params.get("season", None)
        player_id = self.request.query_params.get("player_id", None)

        if season is not None:
            queryset = queryset.filter(season=season)
        if player_id is not None:
            queryset = queryset.filter(player=player_id)
            queryset = queryset.order_by("-season")

        return queryset


class SeasonSettingsViewSet(viewsets.ModelViewSet):
    serializer_class = SeasonSettingsSerializer

    def get_queryset(self):
        queryset = SeasonSettings.objects.all()
        is_active = self.request.query_params.get("is_active", None)
        season = self.request.query_params.get("season", None)

        if season is not None:
            queryset = queryset.filter(season=season)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)

        return queryset.order_by("-season")


class TokenCreateView(djoser.views.TokenCreateView):
    def _action(self, serializer):
        token = utils.login_user(self.request, serializer.user)
        token_serializer_class = settings.SERIALIZERS.token

        response = Response()
        data = token_serializer_class(token).data

        response.set_cookie(
            key = "access_token",
            path = "/",
            value = data["auth_token"],
            max_age = timedelta(days=30),
            secure = not is_localhost,
            httponly = True,
            samesite = "Lax",
            domain = "api.bhmc.org" if not is_localhost else None,
        )

        response.data = "Welcome!"
        response.status_code = status.HTTP_200_OK
        return response


class TokenDestroyView(djoser.views.TokenDestroyView):
    """Use this endpoint to logout user (remove user authentication token)."""

    permission_classes = settings.PERMISSIONS.token_destroy

    def post(self, request):
        response = Response()
        response.delete_cookie(
            key = "access_token",
            path = "/",
            samesite = "Lax",
            domain = "api.bhmc.org" if not is_localhost else None,
        )
        response.status_code = status.HTTP_204_NO_CONTENT
        utils.logout_user(request)
        return response


@api_view(("POST",))
@permission_classes((permissions.IsAuthenticated,))
def import_champions(request):

    event_id = request.data.get("event_id", 0)
    document_id = request.data.get("document_id", 0)

    event = Event.objects.get(pk=event_id)
    document = Document.objects.get(pk=document_id)
    players = Player.objects.all()
    player_map = {player.player_name(): player for player in players}
    existing_champions = {champ.player.id: champ for champ in list(MajorChampion.objects.filter(event=event))}
    season = event.season
    event_name = event.name
    failures = []

    with document.file.open("r") as content:
        wb = load_workbook(filename=content.path, read_only=True)
        sheet = wb.active
        last_row = sheet.max_row

        # skip header row
        for i in range(2, last_row + 1):
            flight = sheet.cell(row=i, column=1).value
            if flight is None:
                break

            champion = sheet.cell(row=i, column=2).value
            score = sheet.cell(row=i, column=3).value
            is_net = False if sheet.cell(row=i, column=4).value is None else sheet.cell(row=i, column=4).value

            try:
                players, errors = get_players(champion, player_map)
                team_id = "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
                for player in players:
                    existing = existing_champions.get(player.id)
                    if existing is None:
                        new_champion = MajorChampion.objects.create(season=season,
                                                                    event=event,
                                                                    event_name=event_name,
                                                                    flight=flight,
                                                                    player=player,
                                                                    team_id=team_id,
                                                                    score=score,
                                                                    is_net=is_net)
                        new_champion.save()
                    else:
                        existing.flight = flight
                        existing.score = score
                        existing.is_net = is_net
                        existing.team_id = team_id
                        existing.save()

                for error in errors:
                    failures.append(error)

            except Exception as ex:
                failures.append(str(ex))

    # do not keep the data file
    document.file.delete()
    document.delete()

    return Response(data=failures, status=200)


def get_players(champion, player_map):
    players = []
    errors = []
    partners = champion.split(" + ")
    for partner in partners:
        player = player_map.get(partner.strip())
        if player is None:
            errors.append(f"Player {partner} not found")
        else:
            players.append(player)

    return players, errors
